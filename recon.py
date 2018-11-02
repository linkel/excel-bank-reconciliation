#!/usr/bin/python3

import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
import sys
import time
import os
import glob

os.chdir(".")       # bind path to the current directory

#highlight cells in openpyxl
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=8)
#bd = Side(style='thick', color="000000")
#highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = PatternFill(fill_type='lightUp',
                 start_color='fff000',
                end_color='6efdfd')

red = NamedStyle(name="red")
red.font = Font(bold=True, size=8)
#bd = Side(style='thick', color="000000")
#red.border = Border(left=bd, top=bd, right=bd, bottom=bd)
red.fill = PatternFill(fill_type='lightUp',
                 start_color='fff000',
                end_color='FD6E6E')

print("\n")
print("files available in this folder:")
print("\n")

# enumerate files in current directory
files = os.listdir()
i = 1
for f in glob.glob("*.xlsx"):
    print("(" + str(i) + "). "  + str(f))
    i+= 1

print("\n")
print("enter the name of your excel file without extension:")
wb_name = input()
print("\n")
print("opening your book...")
try:
    workBook = openpyxl.load_workbook(wb_name + str(".xlsx"))
except IOError:
    print("could not find the book. exiting...")
    sys.exit()
print("I found the following sheets in your book:")
print("\n")
j = 1
for sheets in workBook.sheetnames:
    print("(" + str(j) + "). "  + str(sheets))
    j += 1
print("\n")
print("enter the sheet name with data to compare:")
b_sheet = input()
print("\n")
try:
    bankSheet = workBook.get_sheet_by_name(b_sheet)
except KeyError:
    print("no such sheet in your file: " + wb_name + " xlsx.")
    print("exitting....")
    sys.exit()

print("SUCCESS: data found at sheet: " + b_sheet)
print("\n")
print("enter the sheet name with second data to compare:")
u_sheet = input()
try:
    userSheet = workBook.get_sheet_by_name(u_sheet)
    print("\n")
    print("SUCCESS: data found at sheet: " + u_sheet)
    print("\n")
except KeyError:
    print("no such sheet found in your file: " + wb_name + " xlsx.")
    sys.exit()

def get_amount(sheetName):
    amount = []
    for row in range(2, sheetName.max_row + 1):
        cellObj = sheetName["I" + str(row)]
        eachAmu = cellObj.value
        if eachAmu != '' and eachAmu != None and eachAmu != 0:
            eachAmu = abs(float(eachAmu))
            amount.append(round(eachAmu,2))
    return amount

def get_last5digits(sheetName):
    vnumbers = []
    for row in range(2, sheetName.max_row + 1):
        cellObj = sheetName["P" + str(row)]
        eachAmu = cellObj.value
        if eachAmu != '' and eachAmu != None and eachAmu != 0:
            eachAmu = eachAmu[-6:]
            if eachAmu[0] == "V":
                vnumbers.append(eachAmu[1:])
    return vnumbers

amounts = get_amount(bankSheet)
vnumbers = get_last5digits(bankSheet)

print("Number of total amounts found: " + str(len(amounts)))
print("\n")

print("processing your data....")
print("finding matches...")
count = 0    # keep track of matches found
matches = [] # another tracking of matches to go back to highlight the bank statement sheet
for row in range(2, userSheet.max_row + 1):
    AmcellObject = userSheet["J" + str(row)]        # same as above for every record on column J of the excel file
    if AmcellObject.value != '' and AmcellObject.value != None and AmcellObject.value != 0 and isinstance(AmcellObject.value, str) == False:
        if abs(AmcellObject.value) in amounts:           #check for matches in the "amount" column
            AmcellObject.style = highlight          
            matches.append(abs(AmcellObject.value))
            amounts.remove(abs(AmcellObject.value))
            count += 1
        elif AmcellObject.value != '' and AmcellObject.value != None and AmcellObject.value != 0:
            AmcellObject.style = red
print(str(count) + " matches found")
print("\n")  

# vnumber possible ID

print("finding possible account number matches...")
newcount = 0   # keep track of account matches found
sumCount = 0   # keep track of sums added
acctmatches = [] # another tracking of matches to go back to highlight the bank statement sheet
for row in range(2, userSheet.max_row + 1):
    AmcellObject = userSheet["H" + str(row)]        #same as above for every record on column H of the excel file
    eachObj = str(AmcellObject.value)
    if eachObj != '' and eachObj != None and eachObj != 0:
        eachObj = eachObj[-5:]
        if eachObj in vnumbers:           #check for matches in the "account" column
            thisNumber = eachObj
            sumList = []
            objectList = []
            try:
                for nextfew in range(row, row + 8):
                    VnumberObject = userSheet["H" + str(nextfew)]
                    Transaction = userSheet["J" + str(nextfew)]
                    indiv = str(VnumberObject.value)
                    if indiv != '' and indiv != None and indiv != 0:
                        indiv = indiv[-5:]
                        if indiv == thisNumber and Transaction.value != None and Transaction.value != 0:
                            sumList.append(abs(float(Transaction.value)))
                            objectList.append(Transaction)
                if sum(sumList) in amounts:
                    sumCount += 1
                    for stuff in objectList:
                        stuff.style = highlight
            except:
                print("Reached end of line while checking.")
            AmcellObject.style = highlight          
            acctmatches.append(eachObj)
            newcount += 1
print(str(newcount) + " possible account number matches found")
#print(str(sumCount) + " sums added to highlight")

print("\n")
for row in range(2, bankSheet.max_row + 1):
    AmcellObject = bankSheet["P" + str(row)]        # same as above for every record on column P of the excel file
    eachObj = str(AmcellObject.value)
    if eachObj != '' and eachObj != None and eachObj != 0:
        eachObj = eachObj[-5:]
        if eachObj in acctmatches:           # check for matches in the "account" column
            AmcellObject.style = highlight    


# this next part tries to sum nearby numbers with same department numbers...WIP
sums = 0
for row in range(2, userSheet.max_row + 1):
    theCellObj = userSheet["G" + str(row)]
    numberacross = userSheet["J" + str(row)]
    departmentnumber = theCellObj.value
    collection = []
    objcoll = []
    if numberacross.value != None and numberacross.value != "Balance ":
        #print(numberacross.value)
        collection.append(abs(float(numberacross.value)))
        objcoll.append(numberacross)
    for nextrow in range(row+1, row+18):
        nextCellObj = userSheet["G" + str(nextrow)]
        nextnumberacross = userSheet["J"+str(nextrow)]
        #print(nextnumberacross.value)
        if nextCellObj.value != None:
            if nextCellObj.value == departmentnumber:
                if nextnumberacross.value != None and nextnumberacross.value not in matches:
                    collection.append(abs(float(nextnumberacross.value)))
                    objcoll.append(nextnumberacross)
                    #print("Appending")
            if nextCellObj.value != None and nextCellObj.value != departmentnumber:
                #print(str(nextrow) + "break")
                break
    #print(str(row) +" done, current sum: " + str(round(sum(collection),2)))
    if len(collection) > 1 and round(sum(collection),2) in amounts:
        matches.append(round(sum(collection),2))
        for stuff in objcoll:
            stuff.style = highlight
            #print("found collection sums")
            sums += 1

# highlighting the matches in the banksheet
for row in range(2, bankSheet.max_row + 1):
    AmcellObject = bankSheet["I" + str(row)]        # same as above for every record on column I of the excel file
    eachObj = AmcellObject.value
    if eachObj != '' and eachObj != None and eachObj != 0:
        if abs(eachObj) in matches:           #check  for matches in the "amount" column
            AmcellObject.style = highlight
            matches.remove(abs(eachObj)) # I don't think I need to remove for the matches, right? -new note: yes you did.
        else:
            AmcellObject.style = red      

print("SUCCESS:" + str(count) + " transaction matches highlighted")
print("SUCCESS:" + str(newcount) + " possible account matches highlighted")
print("SUCCESS:" + str(sums) + " sums matches highlighted")

print("creating new file in your folder....")
workBook.save("ready.xlsx")             # create new file with all the matched instance highlighted automatically
print("ready.xlsx created")
print("Exiting...")

      


