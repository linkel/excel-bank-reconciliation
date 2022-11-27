#!/usr/bin/python3

import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
import sys
import time
import os
import glob

# This recon python program is to reduce the amount of upfront work
# by trying to find matches between the two excel sheets
# and highlight them yellow for nonmatching values,
# green if matches perfectly,
# and orange if there are no entries in the second sheet

wb_name = sys.argv[1]
alco_sheet_name = sys.argv[2]
fa_sheet_name = sys.argv[3]

# bind path to the current directory
os.chdir(".")

# declare some highlight styles
# green
matched_color = NamedStyle(name="matched_color")
matched_color.font = Font(bold=True, size=8)
matched_color.fill = PatternFill(fill_type='lightUp',
                 start_color='fff000',
                end_color='6efdfd')

# yellow
no_match_color = NamedStyle(name="no_match_color")
no_match_color.font = Font(bold=True, size=8)
no_match_color.fill = PatternFill(fill_type='lightUp',
                 start_color='fff000',
                end_color='FFFF00')

# orange
no_result_color = NamedStyle(name="no_result_color")
no_result_color.font = Font(bold=True, size=8)
no_result_color.fill = PatternFill(fill_type='lightUp',
                 start_color='fff000',
                end_color='FFA500')

wb = openpyxl.load_workbook(wb_name + str(".xlsx"))

alco_sheet = wb.get_sheet_by_name(alco_sheet_name)
fa_sheet = wb.get_sheet_by_name(fa_sheet_name)

# we care about alco first, check column G
# which contains Journal Ids which can be deposit numbers
# or journals
def get_unique_journal_ids_and_sums(sheet_obj):
    ids_to_amt = {}
    for row in range(2, sheet_obj.max_row + 1):
        id_cell = sheet_obj["G" + str(row)]
        id_cell_value = id_cell.value

        transaction_amt_cell = sheet_obj["I" + str(row)]
        transaction_amt = float(transaction_amt_cell.value)
        # Just add DEP and 0000 ones
        # don't bother adding the others yet
        # DEP is a deposit number
        # 0000 is a journal number prefix
        if id_cell_value.startswith('DEP'):
            id_cell_value = id_cell_value[3:]
            # for some reason getting tons of zeroes and a little bit of change at the end so I will round
            ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + transaction_amt, 2)
        elif id_cell_value.startswith('0000'):
            id_cell_value = id_cell_value[4:]
            ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + transaction_amt, 2)
    # print(ids_to_amt)
    return ids_to_amt


def get_fa_journal_id_and_sums(sheet_obj):
    ids_to_amt = {}
    # range 10 to skip the header
    for row in range(10, sheet_obj.max_row + 1):
        id_cell = sheet_obj["G" + str(row)]
        id_cell_value = id_cell.value

        trans_amt_cell = sheet_obj["J" + str(row)]
        trans_amt_cell_value = trans_amt_cell.value
        if (id_cell_value != None and trans_amt_cell_value != None):
            id_cell_value = id_cell_value.lower()
            trans_amt = float(trans_amt_cell_value)

            if id_cell_value.startswith('0000'):
                id_cell_value = id_cell_value[4:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            # bruh
            elif id_cell_value.startswith('dep: jid'):
                id_cell_value = id_cell_value[8:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            elif id_cell_value.startswith('dep: '):
                id_cell_value = id_cell_value[5:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            elif id_cell_value.startswith('dep:'):
                id_cell_value = id_cell_value[4:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            elif id_cell_value.startswith('dep : '):
                id_cell_value = id_cell_value[6:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            elif id_cell_value.startswith('dep :'):
                id_cell_value = id_cell_value[5:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            elif id_cell_value.startswith('dep # '):
                id_cell_value = id_cell_value[6:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            elif id_cell_value.startswith('dep #'):
                id_cell_value = id_cell_value[5:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            elif id_cell_value.startswith('dep '):
                id_cell_value = id_cell_value[4:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            elif id_cell_value.startswith('dep'):
                id_cell_value = id_cell_value[3:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            elif id_cell_value.startswith('jid '):
                id_cell_value = id_cell_value[4:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            elif id_cell_value.startswith('jid'):
                id_cell_value = id_cell_value[3:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            elif id_cell_value.startswith('dp '):
                id_cell_value = id_cell_value[3:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            elif id_cell_value.startswith('dp'):
                id_cell_value = id_cell_value[2:]
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)
            else:
                ids_to_amt[id_cell_value] = round(ids_to_amt.setdefault(id_cell_value, 0) + trans_amt, 2)

    # print(ids_to_amt)
    return ids_to_amt

# mutates cell
def do_color(cell, t_cell, value, good_ids, no_match_ids, no_res_ids):
    if value in good_ids:
        cell.style = matched_color
        t_cell.style = matched_color

    elif value in no_match_ids:
        cell.style = no_match_color
        t_cell.style = no_match_color

    elif value in no_res_ids:
        cell.style = no_result_color
        t_cell.style = no_result_color


def color_alco_rows(sheet_obj, good_ids, no_match_ids, no_res_ids):
    for row in range(2, sheet_obj.max_row + 1):
        cell = sheet_obj["G" + str(row)]
        id_cell_value = cell.value
        t_cell = sheet_obj["I" + str(row)]
        
        if id_cell_value.startswith('DEP'):
            id_cell_value = id_cell_value[3:]
            do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
        elif id_cell_value.startswith('0000'):
            id_cell_value = id_cell_value[4:]
            do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)

def color_fa_rows(sheet_obj, good_ids, no_match_ids, no_res_ids):
    for row in range(10, sheet_obj.max_row + 1):
        cell = sheet_obj["G" + str(row)]
        id_cell_value = cell.value

        # here for filtering
        t_cell = sheet_obj["J" + str(row)]
        t_cell_val = t_cell.value
        
        if (id_cell_value != None and t_cell_val != None):
            id_cell_value = id_cell_value.lower()
            if id_cell_value.startswith('0000'):
                id_cell_value = id_cell_value[4:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('dep: jid'):
                id_cell_value = id_cell_value[8:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('dep: '):
                id_cell_value = id_cell_value[5:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('dep:'):
                id_cell_value = id_cell_value[4:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('dep : '):
                id_cell_value = id_cell_value[6:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('dep :'):
                id_cell_value = id_cell_value[5:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('dep # '):
                id_cell_value = id_cell_value[6:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('dep #'):
                id_cell_value = id_cell_value[5:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('dep '):
                id_cell_value = id_cell_value[4:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('dep'):
                id_cell_value = id_cell_value[3:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('jid '):
                id_cell_value = id_cell_value[4:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('jid'):
                id_cell_value = id_cell_value[3:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('dp '):
                id_cell_value = id_cell_value[3:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            elif id_cell_value.startswith('dp'):
                id_cell_value = id_cell_value[2:]
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)
            else:
                do_color(cell, t_cell, id_cell_value, good_ids, no_match_ids, no_res_ids)



alco_sums = get_unique_journal_ids_and_sums(alco_sheet)

fa_sums = get_fa_journal_id_and_sums(fa_sheet)

ids_to_color_green = set()
ids_to_color_yellow = set()
ids_to_color_orange = set()

for k, v in alco_sums.items():
    if k in fa_sums:
        fa_v = fa_sums[k] 
        if abs(v) == abs(fa_v):
            ids_to_color_green.add(k)
        else:
            ids_to_color_yellow.add(k)
    else:
        ids_to_color_orange.add(k)

color_alco_rows(alco_sheet, ids_to_color_green, ids_to_color_yellow, ids_to_color_orange)
color_fa_rows(fa_sheet, ids_to_color_green, ids_to_color_yellow, ids_to_color_orange)

print(ids_to_color_green)

print(ids_to_color_yellow)

print("Found " + str(len(alco_sums)) + " unique ids in col G from Alcolink")
print("Found " + str(len(fa_sums)) + " unique ids in col G from FA")
print("Colored " + str(len(ids_to_color_green)) + " ids green for exact match")
print("Colored " + str(len(ids_to_color_yellow)) + " ids yellow for no match")
print("Colored " + str(len(ids_to_color_orange)) + " ids orange for no result in FA")

wb.save("ready.xlsx")
print("ready.xlsx created")
print("Exiting...")