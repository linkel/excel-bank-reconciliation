#!/usr/bin/python3

import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
import sys
import time
import os
import glob

os.chdir(".")       # bind path to the current directory

#highlight cells in openpyxl
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=8)
highlight.fill = PatternFill(fill_type='lightUp',
                 start_color='fff000',
                end_color='6efdfd')

red = NamedStyle(name="red")
red.font = Font(bold=True, size=8)
red.fill = PatternFill(fill_type='lightUp',
                 start_color='fff000',
                end_color='FD6E6E')

print("\n")
print("files available in this folder:")
print("\n")

# enumerate files in current directory
files = os.listdir()
i = 1
for f in glob.glob("*.xlsx"):
    print("(" + str(i) + "). "  + str(f))
    i+= 1

print("\n")
print("Enter the name of your excel file without extension:")
wb_name = input()
print("\n")
print("Opening your book...")
try:
    workBook = openpyxl.load_workbook(wb_name + str(".xlsx"))
except IOError:
    print("could not find the book. exiting...")
    sys.exit()
print("Found the following sheets in your book:")
print("\n")
j = 1
for sheets in workBook.sheetnames:
    print("(" + str(j) + "). "  + str(sheets))
    j += 1
print("\n")
print("Enter the sheet name with data to compare:")
b_sheet = input()
print("\n")
try:
    bankSheet = workBook.get_sheet_by_name(b_sheet)
except KeyError:
    print("no such sheet in your file: " + wb_name + " xlsx.")
    print("exiting....")
    sys.exit()

print("SUCCESS: data found at sheet: " + b_sheet)
print("\n")
print("Enter the sheet name with second data to compare:")
u_sheet = input()
try:
    userSheet = workBook.get_sheet_by_name(u_sheet)
    print("\n")
    print("SUCCESS: data found at sheet: " + u_sheet)
    print("\n")
except KeyError:
    print("no such sheet found in your file: " + wb_name + " xlsx.")
    sys.exit()

def get_amount(sheetName):
    amount = []
    for row in range(2, sheetName.max_row + 1):
        cellObj = sheetName["I" + str(row)]
        eachAmu = cellObj.value
        if eachAmu != '' and eachAmu != None and eachAmu != 0 and isinstance(eachAmu, str) == False:
            eachAmu = -(float(eachAmu)) # since Alcolink is negative numbers
            amount.append(round(eachAmu,2))
    return amount

# for voucher numbers, cell P on Alcolink and cell H on front accounting sheet
def get_last5digits(sheetName):
    vnumbers = []
    for row in range(2, sheetName.max_row + 1):
        cellObj = sheetName["P" + str(row)]
        eachAmu = cellObj.value
        if eachAmu != '' and eachAmu != None and eachAmu != 0:
            eachAmu = eachAmu[-6:]
            if eachAmu[0] == "V":
                vnumbers.append(eachAmu[1:])
    return vnumbers

amounts = get_amount(bankSheet) # from Alcolink I
vnumbers = get_last5digits(bankSheet) # from Alcolink P

print("Number of total amounts found: " + str(len(amounts)))
print("\n")

print("Processing your data....")
print("Finding matches...")
count = 0    # keep track of matches found
matches = [] # another tracking of matches to go back to highlight the Alcolink sheet
for row in range(2, userSheet.max_row + 1):
    FA_cellobj = userSheet["J" + str(row)]
    if FA_cellobj.value != '' and FA_cellobj.value != None and FA_cellobj.value != 0 and isinstance(FA_cellobj.value, str) == False:
        if FA_cellobj.value in amounts:           #check for matches in the "amount" column
            FA_cellobj.style = highlight          
            matches.append(FA_cellobj.value)
            amounts.remove(FA_cellobj.value)
            count += 1
        else:
            FA_cellobj.style = red
print(str(count) + " matches found")
print("\n")  

# vnumber possible ID

print("finding possible account number matches...")
newcount = 0   # keep track of account matches found
acctmatches = [] # another tracking of matches to go back to highlight the bank statement sheet
for row in range(2, userSheet.max_row + 1):
    FA_cellobj = userSheet["H" + str(row)]        #same as above for every record on column H of the excel file
    eachObj = str(FA_cellobj.value)
    if eachObj != '' and eachObj != None and eachObj != 0:
        eachObj = eachObj[-5:]
        if eachObj in vnumbers:
            FA_cellobj.style = highlight          
            acctmatches.append(eachObj)
            newcount += 1
print(str(newcount) + " possible account number matches found")


# Vnumber match highlight in Alcolink Sheet
print("\n")
for row in range(2, bankSheet.max_row + 1):
    AlcoCellObj = bankSheet["P" + str(row)]        # same as above for every record on column P of the excel file
    eachObj = str(AlcoCellObj.value)
    if eachObj != '' and eachObj != None and eachObj != 0:
        eachObj = eachObj[-5:]
        if eachObj in acctmatches:           # check for matches in the "account" column
            AlcoCellObj.style = highlight    
            acctmatches.remove(eachObj)

# highlighting the matches in the Alcolink Sheet
for row in range(2, bankSheet.max_row + 1):
    AlcoCellObj = bankSheet["I" + str(row)]        # same as above for every record on column I of the excel file
    value = AlcoCellObj.value
    if value != '' and value != None and value != 0 and isinstance(value, str) == False:
        if -value in matches:           #check for matches in the "matches" list
            AlcoCellObj.style = highlight
            matches.remove(-value)
        else:
            AlcoCellObj.style = red      

print("SUCCESS:" + str(count) + " transaction matches highlighted")
print("SUCCESS:" + str(newcount) + " possible account matches highlighted")
print("creating new file in your folder....")
workBook.save("ready.xlsx")             # create new file with all the matched instance highlighted automatically
print("ready.xlsx created")
print("Exiting...")

      


